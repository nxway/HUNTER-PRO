"""
inbox_import.py — обратный импорт заполненного файла из inbox/ (разделы
8.4 и 7.5 ТЗ). Необязателен: не положил файл — ничего не сломается,
следующий запуск просто найдёт inbox/ пустым.

Читает по каждой строке колонки "ИНН", "Итог", "Заметка". По "Итогу"
(один из пяти исходов дропдауна, см. _OUTCOME_RULES) записывает touches,
двигает companies.my_status и companies.note — единственные два поля,
которые этот модуль трогает в companies напрямую (это и есть его работа,
в отличие от enrich/*, которому эти поля трогать запрещено правилом 4.3.2).

Правило возврата (7.5 ТЗ): исходы 2-4 (возит сам / есть экспедитор / не тот
профиль) — компания может снова попасть в выгрузку через 180 дней, исход 5
(не дозвонился) — через 7 дней. Дата возврата пишется в touches.next_date;
export.py учитывает её в _candidates() наравне с правилом "новый сигнал +
REEXPORT_AFTER_DAYS" из этапа 3 — оба пути к повторному всплытию равноправны.

Массовое применение вердикта (7.5 ТЗ): "не тот профиль" понижает в жёлтую
корзину похожие непрозвоненные компании (тот же код продукции, тот же
источник сигнала) — не в красную, это не доказательство мусора, только
слабый сигнал (7.1, 7.3 ТЗ).

Обработанные файлы переносятся в inbox/processed/, чтобы не читать их
повторно при следующем запуске.

Запуск: python inbox_import.py
"""
from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl

import config
import db

_OUTCOME_RULES = {
    "дал заявку": {"my_status": "client", "reactivate_days": None, "mass_verdict": False},
    "возит сам": {"my_status": "refused", "reactivate_days": 180, "mass_verdict": False},
    "есть экспедитор": {"my_status": "refused", "reactivate_days": 180, "mass_verdict": False},
    "не тот профиль": {"my_status": "refused", "reactivate_days": 180, "mass_verdict": True},
    "не дозвонился": {"my_status": "called", "reactivate_days": 7, "mass_verdict": False},
}


def _apply_mass_verdict(conn, source_inn: str) -> int:
    """"Не тот профиль" -> похожие непрозвоненные (тот же код продукции,
    тот же источник) понижаются из зелёной в жёлтую корзину (7.5 ТЗ)."""
    row = conn.execute("SELECT product_code FROM companies WHERE inn = ?", (source_inn,)).fetchone()
    if not row or not row["product_code"]:
        return 0

    src_row = conn.execute(
        "SELECT source FROM signals WHERE inn = ? ORDER BY signal_date DESC LIMIT 1",
        (source_inn,),
    ).fetchone()

    query = """
        UPDATE companies
        SET bucket = 'yellow',
            bucket_reason = 'похожа на компанию с отказом "не тот профиль"'
        WHERE product_code = :product_code
          AND my_status = 'new'
          AND bucket = 'green'
          AND inn != :inn
    """
    params = {"product_code": row["product_code"], "inn": source_inn}
    if src_row and src_row["source"]:
        query += " AND inn IN (SELECT inn FROM signals WHERE source = :source)"
        params["source"] = src_row["source"]

    cur = conn.execute(query, params)
    return cur.rowcount


def import_file(conn, path: Path) -> dict:
    """Разбирает один заполненный .xlsx. Не находит ожидаемых колонок в
    шапке — явная ошибка (файл не тот/повреждён), а не тихий пропуск."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        col = {name: headers.index(name) for name in ("ИНН", "Итог", "Заметка")}
    except ValueError as exc:
        raise ValueError(f"{path}: не нашёл ожидаемые колонки в шапке {headers}") from exc

    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    today = date.today()
    stats = {"processed": 0, "unknown_outcome": 0, "unknown_inn": 0, "mass_downgraded": 0}

    for row in ws.iter_rows(min_row=2):
        inn_cell = row[col["ИНН"]].value
        inn = str(inn_cell).strip() if inn_cell else None
        outcome_cell = row[col["Итог"]].value
        outcome = outcome_cell.strip() if isinstance(outcome_cell, str) else None
        note = row[col["Заметка"]].value

        if not inn or not outcome:
            continue

        if not conn.execute("SELECT 1 FROM companies WHERE inn = ?", (inn,)).fetchone():
            stats["unknown_inn"] += 1
            continue

        rule = _OUTCOME_RULES.get(outcome)
        if not rule:
            stats["unknown_outcome"] += 1
            continue

        next_date = (
            (today + timedelta(days=rule["reactivate_days"])).isoformat()
            if rule["reactivate_days"]
            else None
        )
        conn.execute(
            "INSERT INTO touches (inn, touch_date, result, next_date, created_at) VALUES (?, ?, ?, ?, ?)",
            (inn, today.isoformat(), outcome, next_date, now),
        )

        fields = {"my_status": rule["my_status"], "updated_at": now}
        if note:
            fields["note"] = note
        set_clause = ", ".join(f"{k} = :{k}" for k in fields)
        conn.execute(f"UPDATE companies SET {set_clause} WHERE inn = :inn", {**fields, "inn": inn})

        if rule["mass_verdict"]:
            stats["mass_downgraded"] += _apply_mass_verdict(conn, inn)

        stats["processed"] += 1

    conn.commit()
    return stats


def import_inbox(conn) -> dict:
    """Обходит inbox/*.xlsx, обрабатывает и переносит в inbox/processed/."""
    inbox_dir = config.BASE_DIR / "inbox"
    processed_dir = inbox_dir / "processed"
    inbox_dir.mkdir(parents=True, exist_ok=True)

    totals = {"files": 0, "processed": 0, "unknown_outcome": 0, "unknown_inn": 0, "mass_downgraded": 0}
    for path in sorted(inbox_dir.glob("*.xlsx")):
        stats = import_file(conn, path)
        for key in ("processed", "unknown_outcome", "unknown_inn", "mass_downgraded"):
            totals[key] += stats[key]
        totals["files"] += 1

        processed_dir.mkdir(parents=True, exist_ok=True)
        path.rename(processed_dir / f"{date.today().isoformat()}-{path.name}")

    return totals


def main() -> None:
    from rich.console import Console

    console = Console()
    conn = db.init_db()
    totals = import_inbox(conn)
    conn.close()

    if totals["files"] == 0:
        console.print("inbox_import: файлов в inbox/ нет — это нормально, импорт необязателен (8.4 ТЗ)")
    else:
        console.print(
            f"inbox_import: файлов {totals['files']} · обработано строк {totals['processed']} "
            f"· неизвестный исход {totals['unknown_outcome']} · неизвестный ИНН {totals['unknown_inn']} "
            f"· понижено по массовому вердикту {totals['mass_downgraded']}"
        )


if __name__ == "__main__":
    main()
