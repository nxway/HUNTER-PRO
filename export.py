"""
export.py — сборка .xlsx из базы (раздел VIII ТЗ). Ничего не считает и
никуда не ходит — только читает готовое из companies/signals и пишет файл
(правило слоёв 2.1 ТЗ). Красная корзина в файл не попадает вообще (8.2).

Правило 8.3 ТЗ, главное для всего проекта: одну компанию выгружаем один
раз. exported_at проставляется в базе только ПОСЛЕ успешной записи файла
на диск, в той же функции — не записался файл, не проставится дата
(раздел "Грабли" ТЗ).

Запуск: python export.py
"""
from __future__ import annotations

import json
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config
import db

_BUCKET_LABELS = {"green": "зелёная", "yellow": "жёлтая"}
_SIGNAL_LABELS = {"new_declaration": "декларация"}
_OUTCOME_CHOICES = ["дал заявку", "возит сам", "есть экспедитор", "не тот профиль", "не дозвонился"]

_HEADERS = [
    "Компания", "Телефон", "Город", "Что возит", "Кузов",
    "Почему в списке", "Риск", "Корзина", "ИНН", "Итог", "Заметка",
]
_COL = {name: i + 1 for i, name in enumerate(_HEADERS)}

_YELLOW_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
_COLUMN_WIDTHS = [32, 16, 16, 20, 10, 24, 14, 10, 14, 16, 30]


def _candidates(conn) -> list[dict]:
    """Компании green/yellow, которые ещё не выгружались, либо всплывают
    повторно одним из двух равноправных путей:
      - новый сигнал после прошлой выгрузки И прошло больше
        REEXPORT_AFTER_DAYS дней (8.3 ТЗ);
      - подошёл срок возврата после исхода звонка (touches.next_date,
        7.5/8.4 ТЗ: полгода для "возит сам"/"есть экспедитор"/"не тот
        профиль", неделя для "не дозвонился") — и после этого исхода ещё
        не было новой выгрузки (иначе один и тот же возврат всплывал бы
        в файле бесконечно)."""
    rows = conn.execute(
        """
        SELECT c.*, (
            SELECT s.summary || '||' || s.signal_date || '||' || s.type
            FROM signals s WHERE s.inn = c.inn
            ORDER BY s.signal_date DESC LIMIT 1
        ) AS last_signal
        FROM companies c
        WHERE c.bucket IN ('green', 'yellow')
          AND (
              c.exported_at IS NULL
              OR (
                  julianday('now') - julianday(c.exported_at) > :reexport_days
                  AND EXISTS (
                      SELECT 1 FROM signals s
                      WHERE s.inn = c.inn AND s.created_at > c.exported_at
                  )
              )
              OR EXISTS (
                  SELECT 1 FROM touches t
                  WHERE t.inn = c.inn
                    AND t.next_date <= date('now')
                    AND (c.exported_at IS NULL OR t.created_at > c.exported_at)
              )
          )
        ORDER BY c.bucket, c.name
        """,
        {"reexport_days": config.REEXPORT_AFTER_DAYS},
    ).fetchall()
    return [dict(r) for r in rows]


def _suggestion_rows(conn) -> list[dict]:
    """Раздел 7.5 ТЗ: "правила система предлагает сама, с цифрами" —
    товарные группы, где много звонков, ноль заявок и много "возит сам".
    Ничего не исключает автоматически — только показывает лист, решение
    вписываешь в конфиг сам."""
    rows = conn.execute(
        """
        SELECT c.product_code AS product_code,
               COUNT(*) AS calls,
               SUM(CASE WHEN t.result = 'дал заявку' THEN 1 ELSE 0 END) AS deals,
               SUM(CASE WHEN t.result = 'возит сам' THEN 1 ELSE 0 END) AS self_haul
        FROM touches t
        JOIN companies c ON c.inn = t.inn
        WHERE c.product_code IS NOT NULL
        GROUP BY c.product_code
        HAVING calls >= 5 AND deals = 0
        ORDER BY calls DESC
        """
    ).fetchall()
    return [dict(r) for r in rows]


def _reason(row: dict) -> str:
    last = row.get("last_signal")
    if not last:
        return ""
    summary, signal_date, signal_type = last.split("||")
    label = _SIGNAL_LABELS.get(signal_type, signal_type)
    try:
        date_str = datetime.fromisoformat(signal_date).strftime("%d.%m")
    except ValueError:
        date_str = signal_date
    return f"{label} от {date_str}"


def _risk_text(raw: Optional[str]) -> str:
    """risk_flags хранится в companies как JSON-список (enrich/risk.py) —
    здесь превращаем в человекочитаемую строку для колонки "Риск" (8.2 ТЗ)."""
    if not raw:
        return ""
    try:
        flags = json.loads(raw)
    except (TypeError, ValueError):
        return str(raw)
    return "; ".join(flags) if isinstance(flags, list) else str(flags)


def build_workbook(rows: list[dict], suggestions: Optional[list[dict]] = None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лиды"
    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append(
            [
                row["name"],
                row.get("phone") or "",
                row.get("city") or "",
                row.get("cargo") or "",
                row.get("body_type") or "",
                _reason(row),
                _risk_text(row.get("risk_flags")),
                _BUCKET_LABELS.get(row["bucket"], row["bucket"]),
                row["inn"],
                "",
                "",
            ]
        )

    # Телефон и ИНН — текстом с текстовым форматом ячейки, иначе Excel
    # испортит ведущий плюс/нули и превратит длинный ИНН в экспоненту
    # (раздел XIII "Грабли" ТЗ).
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=_COL["Телефон"]).number_format = "@"
        ws.cell(row=r, column=_COL["ИНН"]).number_format = "@"
        if ws.cell(row=r, column=_COL["Корзина"]).value == _BUCKET_LABELS["yellow"]:
            for c in range(1, len(_HEADERS) + 1):
                ws.cell(row=r, column=c).fill = _YELLOW_FILL

    for i, width in enumerate(_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.auto_filter.ref = ws.dimensions

    outcome_letter = get_column_letter(_COL["Итог"])
    dv = DataValidation(type="list", formula1='"' + ",".join(_OUTCOME_CHOICES) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{outcome_letter}2:{outcome_letter}{ws.max_row}")

    if suggestions:
        ws2 = wb.create_sheet("Правила")
        ws2.append(["Код продукции", "Звонков", "Заявок", "Возят сами", "Предложение"])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for s in suggestions:
            ws2.append(
                [
                    s["product_code"],
                    s["calls"],
                    s["deals"],
                    s["self_haul"],
                    f"{s['calls']} звонков, {s['deals']} заявок, {s['self_haul']} раз возят сами — исключить?",
                ]
            )
        for i, width in enumerate([16, 10, 10, 12, 60], start=1):
            ws2.column_dimensions[get_column_letter(i)].width = width
        ws2.freeze_panes = "A2"

    return wb


def export_leads(conn) -> Optional[Path]:
    """Собирает .xlsx из непроэкспортированных компаний. Возвращает путь
    к файлу или None, если выгружать нечего."""
    rows = _candidates(conn)
    if not rows:
        return None

    out_dir = config.BASE_DIR / "exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"hunter-{date.today().isoformat()}.xlsx"

    wb = build_workbook(rows, suggestions=_suggestion_rows(conn))
    wb.save(out_path)  # файл на диске — точка невозврата, дальше можно метить exported_at

    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    conn.executemany(
        "UPDATE companies SET exported_at = ?, export_count = COALESCE(export_count, 0) + 1 WHERE inn = ?",
        [(now, row["inn"]) for row in rows],
    )
    conn.commit()

    return out_path


def main() -> None:
    from rich.console import Console

    console = Console()
    conn = db.init_db()
    path = export_leads(conn)
    conn.close()
    if path:
        console.print(f"export: выгружено {path}")
    else:
        console.print("export: новых компаний для выгрузки нет")


if __name__ == "__main__":
    main()
